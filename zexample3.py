import os
import openpyxl

# create output folder if it doesn't exist
if not os.path.exists("output"):
    os.mkdir("output")

# load workbook
workbook = openpyxl.load_workbook("data.xlsx")

# iterate through worksheets and save each as a new workbook
for worksheet in workbook.worksheets:
    new_workbook = openpyxl.Workbook()
    new_workbook.active.title = worksheet.title
    for row in worksheet.iter_rows(values_only=True):
        new_workbook.active.append(row)
    new_workbook.save(f"output/{worksheet.title}.xlsx")
