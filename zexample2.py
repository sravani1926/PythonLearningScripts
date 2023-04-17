import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


# Create output folder if it doesn't exist
if not os.path.exists('output'):
    os.makedirs('output')

# Load the workbook
workbook = load_workbook(filename='data.xlsx', read_only=True)

# Iterate through each sheet in the workbook
for sheet in workbook:
    # Create a new workbook for the sheet
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Copy the values from the original sheet to the new sheet
    for row in sheet.iter_rows(values_only=True):
        new_sheet.append(row)

    # Copy the formatting from the original sheet to the new sheet
    for row in sheet.iter_rows():
        for cell in row:
            column_letter = get_column_letter(cell.column)
            new_cell = new_sheet[f'{column_letter}{cell.row}']
            new_cell._style = cell._style

    # Save the new workbook with the same name as the sheet in the output folder
    new_workbook.save(f'output/{sheet.title}.xlsx')
